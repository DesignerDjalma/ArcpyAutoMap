
import datetime
import os
import string
from openpyxl import load_workbook # type: ignore

def exportaPlanilhaExcel(dpt, messages, caminho=None):
    messages.addMessage("DPT")
    messages.addMessage(dpt)

    def sheetLen(sheet):
        counterLen = 0
        for _ in sheet.rows:
            counterLen += 1
        return counterLen

    def retornaCelulas(sheet):
        return [
            "{}{}".format(col, sheetLen(sheet) + 1)
            for col in string.ascii_uppercase[1:12]
        ]

    def carregarExcel(caminho=None):
        if not caminho:
            modelo = r"\components\excel\modelo.xlsx"
            root_folder = "\\".join(__file__.split("\\")[:-1])
            caminho = root_folder + modelo
        return caminho

    def carregarExcelControle(caminho=None):
        if not caminho:
            modelo = r"\components\excel\CONTROLE.xlsx"
            root_folder = "\\".join(__file__.split("\\")[:-1])
            caminho = root_folder + modelo
        return caminho

    def carregarSaida(pasta_resultado):
        pasta_excel = pasta_resultado + "\\EXCEL"
        if not os.path.exists(pasta_excel):
            os.mkdir(pasta_excel)
        return pasta_excel + "\\Planilha.xlsx"

    # Aqui vão ficar os valores a serem acrescidos
    params = [
        "municipio",
        "situacao",
        "ano",
        "numero",
        "interessado",
        "denominacao",
        "parcela",
        "georreferenciamento",
        "data",
        "complemento",
        "carta",
        "zoneamento",
    ]
    # messages.addMessage("Executando Exportar Planilha Excel")
    dpt["data"] = datetime.datetime.now().strftime("%d/%m/%Y")

    xlsx = carregarExcel(caminho)
    xlsx_saida = carregarSaida(dpt["pasta_resultados"])
    workbook = load_workbook(filename=xlsx)
    sheet = workbook.active
    celulas = retornaCelulas(sheet)

    for cel, param in zip(celulas, params):
        if param in ["situacao", "zoneamento"]:
            valor = dpt[param]
        else:
            valor = dpt[param]
        sheet[cel] = valor # type: ignore
    workbook.save(xlsx_saida)

    # CONTROLE
    xlsxControle = carregarExcelControle(caminho)
    xlsx_saida = carregarSaida(dpt["pasta_resultados"])
    workbookControle = load_workbook(filename=xlsxControle)
    sheetControle = workbookControle.active
    celulasControle = retornaCelulas(sheetControle)

    for cel, param in zip(celulasControle, params):
        if param in ["situacao", "zoneamento"]:
            valor = dpt[param]
        else:
            valor = dpt[param]
        sheetControle[cel] = valor # type: ignore
    workbookControle.save(xlsxControle)
    # messages.addMessage("Finalizado Exportar Planilha Excel")
