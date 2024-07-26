from openpyxl import load_workbook


def adicionar_linha(tabela, dados):
    """
    Adiciona uma linha de dados à tabela.

    Args:
    tabela (str): Caminho para o arquivo da tabela do Excel.
    dados (dict): Dicionário contendo os valores das colunas.
    """
    # Carrega a planilha
    wb = load_workbook(tabela)
    ws = wb.active

    # Encontra a próxima linha vazia
    next_row = ws.max_row + 1

    # Define a ordem das colunas
    colunas = [
        "ID",
        "MUNICIPIO",
        "SITUACAO",
        "ANO",
        "NUMERO",
        "INTERESSADO",
        "IMOVEL",
        "PARCELA",
        "GEORREF",
        "DATA",
        "COMPLEMENTO",
        "CARTA",
        "ZEE",
        "PATH",
    ]

    # Preenche a linha com os dados do dicionário
    for col, header in enumerate(colunas, start=1):
        ws.cell(row=next_row, column=col, value=dados.get(header))

    # Salva as alterações no arquivo
    wb.save(tabela)


# Exemplo de uso
dados_exemplo = {
    "ID": 1,
    "MUNICIPIO": "São Paulo",
    "SITUACAO": "Ativo",
    "ANO": 2024,
    "NUMERO": 123,
    "INTERESSADO": "João Silva",
    "IMOVEL": "Casa",
    "PARCELA": "A",
    "GEORREF": "Georreferenciamento",
    "DATA": "2024-07-26",
    "COMPLEMENTO": "Complemento",
    "CARTA": "Carta",
    "ZEE": "Zona Econômica Especial",
    "PATH": "/path/to/file",
}

adicionar_linha(r"C:\pytscript\tests\PYTESTS\controle.xlsx", dados_exemplo)

